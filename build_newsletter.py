#!/usr/bin/env python3
"""Build a reusable newsletter HTML file from Excel or Google Sheets."""

from __future__ import annotations

import argparse
import csv
import html
import math
import re
import shutil
import sys
from dataclasses import dataclass
from datetime import datetime
from io import StringIO
from pathlib import Path
from urllib.error import HTTPError, URLError
from urllib.parse import quote
from urllib.request import urlopen

from openpyxl import Workbook, load_workbook

MAX_POINTS = 10
NUMBER_PATTERN = re.compile(
    r"(?<!\w)(?:\d{1,3}(?:,\d{3})+|\d+)(?:\.\d+)?(?:[kKmMbBtT%])?(?!\w)"
)
GOOGLE_SHEET_ID_PATTERN = re.compile(r"/spreadsheets/d/([a-zA-Z0-9-_]+)")

DEFAULT_META = {
    "eyebrow": "Globalite Macro Brief",
    "main_title": "WEEKLY TOP 10 ARGUMENTS",
    "subtitle": "A clear weekly macro summary with the key arguments that matter.",
    "block_height": "925000",
    "max_supply_btc": "21000000",
    "circulating_supply_btc": "19960000",
    "hashrate_eh_s": "820",
    "hashrate_scale_eh_s": "1000",
    "snapshot_title": "At The Time Of Writing",
    "snapshot_intro": (
        "At the time of writing, these on-chain supply anchors provide the baseline context."
    ),
    "snapshot_note": "Figures are rounded and updated with each issue.",
    "tldr_title": "TL;DR",
    "tldr_content": (
        "Leverage reset first, liquidity expanded next, and structural adoption kept building."
    ),
    "conclusion_title": "GLOBALITE CONCLUSION",
    "conclusion_content": (
        "For deeper context on these points, visit globalite.co.\n"
        "Our team tracks macro shifts, liquidity, and positioning every week."
    ),
    "cta_url": "https://globalite.co",
    "cta_label": "globalite.co",
    "address_line": "Globalite, Lugano, Piazza dell'Indipendenza 3, CAP 6901",
    "footer_line": "Globalite Macro Brief - For internal distribution.",
    "footer_logo_url": "public/logotosite.png",
    "footer_instagram_icon": "public/instagram.png",
    "footer_x_icon": "public/x:twitter.png",
    "footer_linkedin_icon": "public/linkedin.png",
    "image_dir": ".",
    "auto_image_by_order": "true",
}

DEFAULT_DISTRIBUTION = [
    ("Individuals", 13660000, "rgb(255, 66, 2)"),
    ("Lost Bitcoin", 1570000, "rgb(153, 153, 153)"),
    ("Funds & ETFs", 1490000, "rgb(255, 140, 90)"),
    ("Businesses", 1390000, "rgb(255, 107, 61)"),
    ("To Be Mined", 1040000, "rgb(204, 204, 204)"),
    ("Satoshi / Patoshi", 968000, "rgb(255, 200, 150)"),
    ("Governments", 432000, "rgb(255, 173, 120)"),
    ("Other Entities", 421000, "rgb(255, 227, 180)"),
]


@dataclass
class Point:
    order: int
    title: str
    content: str
    image_path: str
    image_caption: str
    source: str


@dataclass
class DistributionSegment:
    category: str
    amount_btc: float
    percent: float
    color: str


class TabularSheet:
    """Small adapter that exposes CSV rows like an openpyxl worksheet."""

    def __init__(self, rows: list[list[str]]) -> None:
        self.rows = rows

    def iter_rows(
        self,
        min_row: int = 1,
        max_row: int | None = None,
        values_only: bool = False,
    ):
        start = max(min_row - 1, 0)
        stop = max_row if max_row is not None else len(self.rows)
        for row in self.rows[start:stop]:
            # This script only reads values (no cell objects).
            yield tuple(row)


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def parse_bool(value: str) -> bool:
    return normalize_text(value).lower() in {"1", "true", "yes", "y", "on"}


def parse_number(value: object, default: float = 0.0) -> float:
    text = normalize_text(value).replace(",", "")
    if not text:
        return default
    try:
        return float(text)
    except ValueError:
        return default


def parse_order(value: object, row_number: int) -> int:
    text = normalize_text(value)
    if not text:
        raise ValueError(f"Missing order value at points row {row_number}.")
    if isinstance(value, int):
        parsed = value
    elif isinstance(value, float):
        if not value.is_integer():
            raise ValueError(
                f"Order value must be a whole number at points row {row_number}."
            )
        parsed = int(value)
    else:
        if not text.isdigit():
            raise ValueError(
                f"Invalid order value '{text}' at points row {row_number}."
            )
        parsed = int(text)

    if parsed < 1 or parsed > MAX_POINTS:
        raise ValueError(
            f"Order value must be between 1 and {MAX_POINTS} at points row {row_number}."
        )
    return parsed


def extract_google_sheet_id(sheet_ref: str) -> str:
    value = normalize_text(sheet_ref)
    if not value:
        raise ValueError("Google Sheet URL or ID is required.")

    match = GOOGLE_SHEET_ID_PATTERN.search(value)
    if match:
        return match.group(1)

    if re.fullmatch(r"[a-zA-Z0-9-_]{20,}", value):
        return value

    raise ValueError(
        "Invalid --google-sheet value. Pass a full Google Sheet URL or a sheet ID."
    )


def normalize_table_rows(rows: list[list[str]]) -> list[list[str]]:
    if not rows:
        return []
    width = max(len(row) for row in rows)
    return [row + [""] * (width - len(row)) for row in rows]


def next_history_workbook_path(history_dir: Path) -> Path:
    timestamp = datetime.now().strftime("%Y-%m-%d_%H%M")
    backup_path = history_dir / f"newsletter_{timestamp}.xlsx"
    if backup_path.exists():
        timestamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
        backup_path = history_dir / f"newsletter_{timestamp}.xlsx"
    return backup_path


def write_google_snapshot_workbook(
    path: Path,
    meta_rows: list[list[str]],
    points_rows: list[list[str]],
    distribution_rows: list[list[str]],
) -> None:
    workbook = Workbook()
    meta_sheet = workbook.active
    meta_sheet.title = "meta"
    for row in meta_rows:
        meta_sheet.append(row)

    points_sheet = workbook.create_sheet("points")
    for row in points_rows:
        points_sheet.append(row)

    if distribution_rows:
        distribution_sheet = workbook.create_sheet("distribution")
        for row in distribution_rows:
            distribution_sheet.append(row)

    workbook.save(path)


def fetch_google_sheet_rows(
    sheet_id: str,
    tab_name: str,
    required: bool = True,
) -> list[list[str]]:
    safe_tab = quote(normalize_text(tab_name), safe="")
    url = (
        f"https://docs.google.com/spreadsheets/d/{sheet_id}/gviz/tq?"
        f"tqx=out:csv&sheet={safe_tab}"
    )
    try:
        with urlopen(url, timeout=30) as response:
            payload = response.read().decode("utf-8-sig")
    except HTTPError as error:
        if not required and error.code in {400, 404}:
            return []
        raise RuntimeError(
            f"Could not load Google Sheet tab '{tab_name}' (HTTP {error.code}). "
            "Confirm sharing is enabled and tab names are correct."
        ) from error
    except URLError as error:
        raise RuntimeError(
            f"Network error while loading Google Sheet tab '{tab_name}': {error.reason}"
        ) from error

    text = payload.strip()
    if not text:
        if required:
            raise ValueError(f"Google Sheet tab '{tab_name}' is empty.")
        return []

    lowered = text.lower()
    if lowered.startswith("<!doctype html") or lowered.startswith("<html"):
        if not required:
            return []
        raise RuntimeError(
            f"Could not read Google Sheet tab '{tab_name}'. "
            "Share the sheet as viewable (at least 'Anyone with the link can view')."
        )
    if "google.visualization.query.setresponse" in lowered and "status\":\"error\"" in lowered:
        if not required:
            return []
        raise RuntimeError(
            f"Google Sheets returned an error for tab '{tab_name}'. "
            "Check that the tab exists and has access permissions."
        )

    rows = normalize_table_rows(list(csv.reader(StringIO(text))))
    if not rows:
        if required:
            raise ValueError(f"Google Sheet tab '{tab_name}' is empty.")
        return []
    return rows


def create_template_workbook(path: Path, force: bool = False) -> None:
    if path.exists() and not force:
        raise FileExistsError(f"Template already exists at: {path}")

    wb = Workbook()
    meta = wb.active
    meta.title = "meta"
    meta.append(["key", "value"])
    for key, value in DEFAULT_META.items():
        meta.append([key, value])

    points = wb.create_sheet("points")
    points.append(
        ["order", "title", "content", "image_path", "image_caption", "source"]
    )
    points.append(
        [
            1,
            "Liquidity stopped tightening",
            "QT pace has slowed materially.\n- Funding stress eased\n- Repo usage normalized",
            "",
            "Optional image caption",
            "Source: Example Research Desk",
        ]
    )
    points.append(
        [
            2,
            "Market leverage reset",
            "A broad deleveraging event removed excess risk without structural damage.",
            "",
            "",
            "",
        ]
    )

    distribution = wb.create_sheet("distribution")
    distribution.append(["category", "amount_btc", "percent", "color"])
    for category, amount_btc, color in DEFAULT_DISTRIBUTION:
        distribution.append([category, amount_btc, "", color])

    wb.save(path)


def read_meta(meta_sheet) -> dict[str, str]:
    meta: dict[str, str] = {}
    for row in meta_sheet.iter_rows(min_row=2, values_only=True):
        key = normalize_text(row[0] if len(row) > 0 else "")
        value = normalize_text(row[1] if len(row) > 1 else "")
        if key:
            meta[key] = value
    for key, value in DEFAULT_META.items():
        meta.setdefault(key, value)
    return meta


def header_index_map(points_sheet) -> dict[str, int]:
    headers = list(points_sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    mapping: dict[str, int] = {}
    for index, header in enumerate(headers):
        key = normalize_text(header).lower()
        if key:
            mapping[key] = index
    required = ["order", "title", "content", "image_path", "image_caption"]
    missing = [name for name in required if name not in mapping]
    if missing:
        raise ValueError(
            "Missing required columns in points sheet: " + ", ".join(missing)
        )
    return mapping


def read_points(points_sheet) -> list[Point]:
    mapping = header_index_map(points_sheet)
    points: list[Point] = []

    for row_number, row in enumerate(
        points_sheet.iter_rows(min_row=2, values_only=True), start=2
    ):
        order_text = normalize_text(row[mapping["order"]])
        title = normalize_text(row[mapping["title"]])
        content = normalize_text(row[mapping["content"]])
        image_path = normalize_text(row[mapping["image_path"]])
        image_caption = normalize_text(row[mapping["image_caption"]])
        source = (
            normalize_text(row[mapping["source"]])
            if "source" in mapping
            else ""
        )

        if not any([order_text, title, content, image_path, image_caption, source]):
            continue

        order = parse_order(order_text, row_number)
        if not title:
            raise ValueError(f"Missing title at points row {row_number}.")
        if not content:
            raise ValueError(f"Missing content at points row {row_number}.")

        points.append(
            Point(
                order=order,
                title=title,
                content=content,
                image_path=image_path,
                image_caption=image_caption,
                source=source,
            )
        )

    if not points:
        raise ValueError("No points found. Add at least 1 row in the points sheet.")

    points.sort(key=lambda item: item.order)
    orders = [item.order for item in points]
    duplicates = sorted({order for order in orders if orders.count(order) > 1})
    if duplicates:
        raise ValueError(
            "Duplicate order values found: " + ", ".join(str(value) for value in duplicates)
        )
    if len(points) > MAX_POINTS:
        raise ValueError(
            f"Found {len(points)} points. Max allowed is {MAX_POINTS}. "
            "Delete or merge entries to keep it to 10 or fewer."
        )

    return points


def default_distribution_segments(max_supply_btc: float) -> list[DistributionSegment]:
    base = [
        DistributionSegment(
            category=category,
            amount_btc=float(amount),
            percent=0.0,
            color=color,
        )
        for category, amount, color in DEFAULT_DISTRIBUTION
    ]
    return finalize_distribution_segments(base, max_supply_btc)


def finalize_distribution_segments(
    segments: list[DistributionSegment], max_supply_btc: float
) -> list[DistributionSegment]:
    if not segments:
        return []
    denominator = max_supply_btc if max_supply_btc > 0 else sum(s.amount_btc for s in segments)
    if denominator <= 0:
        denominator = 1.0
    normalized = [
        DistributionSegment(
            category=s.category,
            amount_btc=s.amount_btc,
            percent=s.percent if s.percent > 0 else (s.amount_btc / denominator) * 100,
            color=s.color,
        )
        for s in segments
    ]
    total_percent = sum(s.percent for s in normalized)
    if total_percent > 0:
        normalized = [
            DistributionSegment(
                category=s.category,
                amount_btc=s.amount_btc,
                percent=(s.percent / total_percent) * 100,
                color=s.color,
            )
            for s in normalized
        ]
    normalized.sort(key=lambda item: item.amount_btc, reverse=True)
    return normalized


def read_distribution(
    distribution_sheet, max_supply_btc: float
) -> list[DistributionSegment]:
    headers = list(distribution_sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    mapping: dict[str, int] = {}
    for index, header in enumerate(headers):
        key = normalize_text(header).lower()
        if key:
            mapping[key] = index
    required = ["category", "amount_btc", "color"]
    missing = [name for name in required if name not in mapping]
    if missing:
        raise ValueError(
            "Missing required columns in distribution sheet: " + ", ".join(missing)
        )

    segments: list[DistributionSegment] = []
    for row in distribution_sheet.iter_rows(min_row=2, values_only=True):
        category = normalize_text(row[mapping["category"]])
        amount_btc = parse_number(row[mapping["amount_btc"]], default=0.0)
        color = normalize_text(row[mapping["color"]]) or "rgb(255, 66, 2)"
        percent = parse_number(row[mapping["percent"]], default=0.0) if "percent" in mapping else 0.0
        if not category and amount_btc <= 0:
            continue
        if not category:
            raise ValueError("Distribution row is missing category.")
        if amount_btc < 0:
            raise ValueError(f"Distribution amount cannot be negative for category '{category}'.")
        segments.append(
            DistributionSegment(
                category=category,
                amount_btc=amount_btc,
                percent=percent,
                color=color,
            )
        )

    if not segments:
        return default_distribution_segments(max_supply_btc)
    return finalize_distribution_segments(segments, max_supply_btc)


def render_content_blocks(raw: str) -> str:
    lines = raw.replace("\r\n", "\n").split("\n")
    blocks: list[str] = []
    list_open = False

    def close_list() -> None:
        nonlocal list_open
        if list_open:
            blocks.append("</ul>")
            list_open = False

    for original_line in lines:
        line = original_line.strip()
        if not line:
            close_list()
            continue

        if line.startswith("- ") or line.startswith("* "):
            item = emphasize_numbers(line[2:].strip())
            if not list_open:
                blocks.append("<ul>")
                list_open = True
            blocks.append(f"<li>{item}</li>")
        else:
            close_list()
            blocks.append(f"<p>{emphasize_numbers(line)}</p>")

    close_list()
    return "\n".join(blocks)


def emphasize_numbers(text: str) -> str:
    output: list[str] = []
    start = 0
    for match in NUMBER_PATTERN.finditer(text):
        output.append(html.escape(text[start : match.start()]))
        output.append(f"<strong>{html.escape(match.group(0))}</strong>")
        start = match.end()
    output.append(html.escape(text[start:]))
    return "".join(output)


def format_btc_integer(value: float) -> str:
    return f"{int(round(value)):,}"


def format_btc_compact(value: float) -> str:
    abs_value = abs(value)
    if abs_value >= 1_000_000:
        rendered = f"{value / 1_000_000:.2f}".rstrip("0").rstrip(".")
        return f"{rendered}M BTC"
    if abs_value >= 1_000:
        rendered = f"{int(round(value / 1_000)):,}"
        return f"{rendered}K BTC"
    return f"{format_btc_integer(value)} BTC"


def format_percent(value: float) -> str:
    return f"{value:.1f}".rstrip("0").rstrip(".") + "%"


def render_block_height(value: str) -> str:
    clean = normalize_text(value)
    if not clean:
        return "n/a"
    numeric = parse_number(clean, default=-1)
    if numeric >= 0:
        return f"{int(round(numeric)):,}"
    return clean


def render_snapshot_section(
    meta: dict[str, str], distribution: list[DistributionSegment]
) -> str:
    snapshot_title = normalize_text(meta.get("snapshot_title", "")) or "At The Time Of Writing"
    snapshot_intro = normalize_text(meta.get("snapshot_intro", ""))
    snapshot_note = normalize_text(meta.get("snapshot_note", ""))
    max_supply_btc = parse_number(meta["max_supply_btc"], default=21_000_000)
    if max_supply_btc <= 0:
        max_supply_btc = 21_000_000
    circulating_btc = parse_number(meta["circulating_supply_btc"], default=0.0)
    if circulating_btc <= 0:
        circulating_btc = max(
            0.0,
            max_supply_btc
            - sum(
                seg.amount_btc
                for seg in distribution
                if seg.category.strip().lower() == "to be mined"
            ),
        )
    circulation_pct = (circulating_btc / max_supply_btc) * 100 if max_supply_btc > 0 else 0
    hashrate_eh_s = parse_number(meta.get("hashrate_eh_s", "0"), default=0.0)
    hashrate_scale_eh_s = parse_number(meta.get("hashrate_scale_eh_s", "1000"), default=1000.0)
    if hashrate_scale_eh_s <= 0:
        hashrate_scale_eh_s = 1000.0
    hashrate_pct = (hashrate_eh_s / hashrate_scale_eh_s) * 100 if hashrate_scale_eh_s > 0 else 0

    bar_segments = "\n".join(
        f'                    <div class="snapshot-bar-segment" style="width:{max(0.0, seg.percent):.6f}%;background:{html.escape(seg.color)};" title="{html.escape(seg.category)}: {html.escape(format_btc_compact(seg.amount_btc))} ({html.escape(format_percent(seg.percent))})"></div>'
        for seg in distribution
    )
    circumference = 2 * math.pi * 45
    consumed = 0.0
    donut_segments: list[str] = []
    for seg in distribution:
        arc = circumference * (max(0.0, seg.percent) / 100.0)
        donut_segments.append(
            (
                f'                      <circle class="snapshot-donut-segment" cx="60" cy="60" r="45" '
                f'stroke="{html.escape(seg.color)}" stroke-dasharray="{arc:.6f} {circumference:.6f}" '
                f'stroke-dashoffset="{-consumed:.6f}">'
                f'<title>{html.escape(seg.category)}: {html.escape(format_btc_compact(seg.amount_btc))} ({html.escape(format_percent(seg.percent))})</title>'
                "</circle>"
            )
        )
        consumed += arc
    donut_svg = "\n".join(donut_segments)
    legend_rows = "\n".join(
        (
            "                    <div class=\"snapshot-legend-item\">"
            f"<span class=\"snapshot-dot\" style=\"background:{html.escape(seg.color)}\"></span>"
            f"<span class=\"snapshot-name\">{html.escape(seg.category)}</span>"
            f"<span class=\"snapshot-value\">{html.escape(format_btc_compact(seg.amount_btc))} ({html.escape(format_percent(seg.percent))})</span>"
            "</div>"
        )
        for seg in distribution
    )
    return f"""
            <tr>
              <td class="section snapshot">
                <h2>{html.escape(snapshot_title)}</h2>
                <p class="snapshot-intro">{html.escape(snapshot_intro)}</p>
                <div class="snapshot-grid">
                  <div class="snapshot-card">
                    <h3>Ownership Distribution</h3>
                    <div class="snapshot-ownership-viz">
                      <svg class="snapshot-donut" viewBox="0 0 120 120" aria-label="Ownership distribution donut chart">
                        <circle cx="60" cy="60" r="45" fill="none" stroke="#ececec" stroke-width="24"></circle>
{donut_svg}
                        <circle cx="60" cy="60" r="30" fill="#ffffff"></circle>
                        <text x="60" y="56" text-anchor="middle" class="snapshot-donut-label">Supply</text>
                        <text x="60" y="72" text-anchor="middle" class="snapshot-donut-value">{html.escape(format_btc_integer(max_supply_btc))}</text>
                      </svg>
                    </div>
                    <div class="snapshot-bar">
{bar_segments}
                    </div>
                    <div class="snapshot-legend">
{legend_rows}
                    </div>
                  </div>
                  <div class="snapshot-card">
                    <h3>Bitcoin In Circulation At Write Time</h3>
                    <p class="snapshot-circ-value">{html.escape(format_btc_integer(circulating_btc))} BTC</p>
                    <div class="snapshot-progress-track">
                      <div class="snapshot-progress-fill" style="width:{max(0.0, min(100.0, circulation_pct)):.6f}%;"></div>
                    </div>
                    <p class="snapshot-circ-note">{html.escape(format_percent(circulation_pct))} of {html.escape(format_btc_integer(max_supply_btc))} BTC max supply</p>
                  </div>
                  <div class="snapshot-card">
                    <h3>Network Hashrate (Daily)</h3>
                    <p class="snapshot-circ-value">{html.escape(format_btc_integer(hashrate_eh_s))} EH/s</p>
                    <div class="snapshot-progress-track">
                      <div class="snapshot-progress-fill" style="width:{max(0.0, min(100.0, hashrate_pct)):.6f}%;"></div>
                    </div>
                    <p class="snapshot-circ-note">{html.escape(format_percent(hashrate_pct))} of {html.escape(format_btc_integer(hashrate_scale_eh_s))} EH/s reference scale</p>
                  </div>
                </div>
                <p class="snapshot-footnote">{html.escape(snapshot_note)}</p>
              </td>
            </tr>
"""


def looks_like_remote_image_source(path: str) -> bool:
    return path.startswith(("http://", "https://", "data:", "cid:"))


def resolve_image_path(point: Point, meta: dict[str, str], output_dir: Path) -> str:
    image_path = normalize_text(point.image_path)
    if image_path:
        candidate = image_path
    elif parse_bool(meta.get("auto_image_by_order", "true")):
        image_dir = normalize_text(meta.get("image_dir", ".")) or "."
        filename_candidates = [
            f"{point.order}.png",
            f"{point.order}.jpg",
            f"{point.order}.jpeg",
            f"{point.order}.webp",
        ]
        candidate = ""
        for filename in filename_candidates:
            relative_candidate = (Path(image_dir) / filename).as_posix()
            if (output_dir / relative_candidate).exists():
                candidate = relative_candidate
                break
        if not candidate:
            return ""
    else:
        return ""

    if looks_like_remote_image_source(candidate):
        return candidate

    candidate_path = Path(candidate)
    if candidate_path.is_absolute():
        return candidate if candidate_path.exists() else ""

    return candidate if (output_dir / candidate).exists() else ""


def resolve_extra_image_paths(
    point: Point, meta: dict[str, str], output_dir: Path
) -> list[str]:
    if not parse_bool(meta.get("auto_image_by_order", "true")):
        return []

    image_dir = normalize_text(meta.get("image_dir", ".")) or "."
    max_extra_images = int(parse_number(meta.get("max_extra_images", "10"), default=10))
    max_extra_images = max(0, min(20, max_extra_images))
    extensions = ["png", "jpg", "jpeg", "webp"]
    sources: list[str] = []

    for index in range(1, max_extra_images + 1):
        for extension in extensions:
            candidate = (Path(image_dir) / f"{point.order}.{index}.{extension}").as_posix()
            if (output_dir / candidate).exists():
                sources.append(candidate)
                break
    return sources


def render_image_block(point: Point, image_src: str) -> str:
    if not image_src:
        return ""
    caption = point.image_caption or point.title
    return (
        '<div class="image">\n'
        f'  <img src="{html.escape(image_src)}" alt="{html.escape(point.title)}">\n'
        f'  <div class="caption">{html.escape(caption)}</div>\n'
        "</div>"
    )


def render_extra_images_block(point: Point, image_sources: list[str]) -> str:
    if not image_sources:
        return ""
    image_tags = "\n".join(
        f'  <img src="{html.escape(src)}" alt="{html.escape(point.title)} - extra {index}">'
        for index, src in enumerate(image_sources, start=1)
    )
    return '<div class="extra-images">\n' + image_tags + "\n</div>"


def render_point(point: Point, meta: dict[str, str], output_dir: Path) -> str:
    parts = [
        "            <tr>",
        '              <td class="section">',
        f"                <h2>{point.order}. {html.escape(point.title)}</h2>",
    ]

    image_src = resolve_image_path(point, meta, output_dir)
    image_block = render_image_block(point, image_src)
    if image_block:
        parts.append(indent_block(image_block, 16))

    extra_image_sources = resolve_extra_image_paths(point, meta, output_dir)
    extra_images_block = render_extra_images_block(point, extra_image_sources)

    parts.append(indent_block(render_content_blocks(point.content), 16))
    if point.source:
        parts.append(
            f'                <p class="point-source">{html.escape(point.source)}</p>'
        )
    if extra_images_block:
        parts.append(indent_block(extra_images_block, 16))
    parts.extend(
        [
            "              </td>",
            "            </tr>",
        ]
    )
    return "\n".join(parts) + "\n"


def indent_block(text: str, spaces: int) -> str:
    if not text:
        return ""
    prefix = " " * spaces
    return "\n".join(prefix + line if line else "" for line in text.splitlines())


def render_html(
    meta: dict[str, str],
    points: list[Point],
    distribution: list[DistributionSegment],
    output_dir: Path,
) -> str:
    title = html.escape(meta["main_title"])
    subtitle = html.escape(meta["subtitle"])
    eyebrow = html.escape(meta["eyebrow"])
    block_height = html.escape(render_block_height(meta["block_height"]))
    tldr_title = html.escape(meta["tldr_title"])
    tldr_content = indent_block(render_content_blocks(meta["tldr_content"]), 16)
    conclusion_title = html.escape(meta["conclusion_title"])
    conclusion_content = indent_block(render_content_blocks(meta["conclusion_content"]), 16)
    cta_url = html.escape(meta["cta_url"], quote=True)
    cta_label = html.escape(meta["cta_label"])
    address_line = html.escape(meta["address_line"])
    footer_line = html.escape(meta["footer_line"])
    footer_logo_url = html.escape(
        normalize_text(meta.get("footer_logo_url", "public/logotosite.png")),
        quote=True,
    )
    footer_instagram_icon = html.escape(
        normalize_text(meta.get("footer_instagram_icon", "public/instagram.png")),
        quote=True,
    )
    footer_x_icon = html.escape(
        normalize_text(meta.get("footer_x_icon", "public/x:twitter.png")),
        quote=True,
    )
    footer_linkedin_icon = html.escape(
        normalize_text(meta.get("footer_linkedin_icon", "public/linkedin.png")),
        quote=True,
    )

    points_html = "".join(render_point(point, meta, output_dir) for point in points)
    snapshot_html = render_snapshot_section(meta, distribution)

    return f"""<!doctype html>
<html lang="en">
  <head>
    <meta charset="utf-8">
    <meta name="viewport" content="width=device-width, initial-scale=1">
    <title>{title} - Globalite Macro Brief</title>
    <link rel="preconnect" href="https://fonts.googleapis.com">
    <link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
    <link href="https://fonts.googleapis.com/css2?family=Poppins:wght@400;500;700&display=swap" rel="stylesheet">
    <style>
      body {{
        margin: 0;
        padding: 0;
        background: #f5f5f5;
        font-family: "Poppins", Arial, sans-serif;
        color: #1f1f1f;
        -webkit-print-color-adjust: exact;
        print-color-adjust: exact;
      }}
      table {{ border-collapse: collapse; }}
      img {{ border: 0; display: block; max-width: 100%; height: auto; }}
      a {{ color: #ff4202; text-decoration: none; }}
      .toolbar {{ width: 100%; max-width: 680px; margin: 0 auto; display: flex; justify-content: flex-end; padding: 12px 0 8px; }}
      .download-pdf-btn {{ border: 1px solid #ff4202; border-radius: 999px; padding: 8px 14px; background: #ffffff; color: #ff4202; font: 600 12px/1 "Poppins", Arial, sans-serif; cursor: pointer; }}
      .download-pdf-btn:hover {{ background: #fff4ef; }}
      .wrapper {{ width: 100%; background: #f5f5f5; padding: 32px 0; }}
      .container {{ width: 680px; max-width: 680px; background: #ffffff; border: 1px solid #e6e6e6; border-radius: 16px; overflow: hidden; }}
      .divider {{ height: 4px; background: #ff4202; line-height: 4px; }}
      .header {{ padding: 28px 32px 18px; }}
      .logo {{ margin: 0 0 16px; text-align: center; }}
      .logo img {{ width: 190px; margin: 0 auto; }}
      .eyebrow {{ color: #ff4202; font-weight: 700; font-size: 12px; letter-spacing: 1px; text-transform: uppercase; margin: 0 0 6px; }}
      h1 {{ margin: 6px 0 6px; font-size: 28px; line-height: 1.2; font-weight: 700; }}
      .subtitle {{ margin: 0; color: #5f5f5f; font-size: 14px; line-height: 1.6; }}
      .block-height {{ margin: 12px 0 0; display: inline-block; font-size: 12px; line-height: 1.4; color: #8f3a1a; background: #fff1eb; border: 1px solid #ffd6c8; border-radius: 999px; padding: 6px 10px; }}
      .section {{ padding: 16px 32px; border-top: 1px solid #f0f0f0; }}
      .section h2 {{ margin: 0 0 20px; font-size: 18px; font-weight: 700; }}
      .section p {{ margin: 0; font-size: 14px; line-height: 1.6; }}
      .section p + p {{ margin-top: 12px; }}
      .section ul {{ margin: 20px 0 20px 18px; padding: 0; font-size: 14px; line-height: 1.6; }}
      .section li {{ margin-bottom: 8px; }}
      .section .point-source {{ margin-top: 14px; font-size: 11px; line-height: 1.5; color: #8a8a8a; }}
      .image {{ margin: 20px 0; }}
      .image img {{ width: 100%; border-radius: 12px; border: 1px solid #e6e6e6; }}
      .caption {{ font-size: 12px; color: #7a7a7a; margin-top: 6px; }}
      .extra-images {{ margin: 14px 0 0; display: grid; gap: 10px; }}
      .extra-images img {{ width: 100%; border-radius: 12px; border: 1px solid #e6e6e6; }}
      .snapshot {{ background: #fcfcfc; }}
      .snapshot-intro {{ margin: 0 0 14px; font-size: 14px; color: #4f4f4f; }}
      .snapshot-grid {{ display: grid; grid-template-columns: 1fr 1fr 1fr; gap: 12px; }}
      .snapshot-card {{ border: 1px solid #ececec; border-radius: 12px; padding: 12px; background: #ffffff; }}
      .snapshot-card h3 {{ margin: 0 0 10px; font-size: 14px; font-weight: 700; color: #1f1f1f; }}
      .snapshot-ownership-viz {{ display: flex; justify-content: center; margin: 0 0 10px; }}
      .snapshot-donut {{ width: 132px; height: 132px; display: block; }}
      .snapshot-donut-segment {{ fill: none; stroke-width: 24; transform: rotate(-90deg); transform-origin: 60px 60px; }}
      .snapshot-donut-label {{ font-size: 10px; fill: #8a8a8a; }}
      .snapshot-donut-value {{ font-size: 10px; fill: #5f5f5f; }}
      .snapshot-bar {{ width: 100%; height: 24px; border: 1px solid #e6e6e6; border-radius: 8px; overflow: hidden; display: flex; }}
      .snapshot-bar-segment {{ height: 100%; min-width: 2px; }}
      .snapshot-legend {{ margin-top: 10px; display: grid; gap: 6px; }}
      .snapshot-legend-item {{ display: grid; grid-template-columns: 12px 1fr auto; align-items: center; gap: 8px; }}
      .snapshot-dot {{ width: 10px; height: 10px; border-radius: 999px; display: inline-block; }}
      .snapshot-name {{ font-size: 12px; color: #3f3f3f; }}
      .snapshot-value {{ font-size: 12px; color: #5f5f5f; white-space: nowrap; }}
      .snapshot-circ-value {{ margin: 0 0 8px; font-size: 22px; font-weight: 700; color: #1f1f1f; }}
      .snapshot-progress-track {{ width: 100%; height: 14px; border: 1px solid #e0e0e0; border-radius: 999px; background: #f0f0f0; overflow: hidden; }}
      .snapshot-progress-fill {{ height: 100%; background: linear-gradient(90deg, #ff4202 0%, #ff8b61 100%); }}
      .snapshot-circ-note {{ margin: 8px 0 0; font-size: 12px; color: #666666; }}
      .snapshot-footnote {{ margin: 12px 0 0; font-size: 12px; color: #7a7a7a; }}
      .tldr {{ background: #fff8ec; border-top: 2px solid #ff4202; }}
      .conclusion {{ background: #fff7f3; border-top: 2px solid #ff4202; }}
      .footer {{ padding: 18px 32px 28px; font-size: 12px; color: #7a7a7a; }}
      .footer p {{ margin: 0; }}
      .footer p + p {{ margin-top: 6px; }}
      .footer-links {{ margin-top: 14px; padding-top: 12px; border-top: 1px solid #ececec; display: flex; align-items: center; justify-content: space-between; gap: 14px; }}
      .footer-logo-link {{ display: inline-flex; }}
      .footer-logo-link img {{ width: 44px; height: 44px; object-fit: cover; border-radius: 20px; border: 1px solid #e0e0e0; }}
      .footer-social {{ display: flex; align-items: center; gap: 10px; }}
      .footer-social a {{ display: inline-flex; }}
      .footer-social img {{ width: 24px; height: 24px; object-fit: contain; border-radius: 6px; }}
      @media (max-width: 720px) {{
        .toolbar {{ padding: 10px 16px 6px; box-sizing: border-box; }}
        .wrapper {{ padding: 16px 0; }}
        .container {{ width: 100%; max-width: 100%; border-radius: 0; }}
        .header, .section, .footer {{ padding: 18px 20px; }}
        .snapshot-grid {{ grid-template-columns: 1fr; }}
        .footer-links {{ flex-direction: column; align-items: flex-start; }}
        h1 {{ font-size: 24px; }}
      }}
      @media print {{
        .no-print {{ display: none !important; }}
        body {{ background: #ffffff; }}
        .wrapper {{ background: #ffffff; padding: 0; }}
        .container {{ border: 0; border-radius: 0; }}
      }}
    </style>
  </head>
  <body>
    <div class="toolbar no-print">
      <button class="download-pdf-btn" type="button" onclick="window.print()">Download PDF</button>
    </div>
    <table class="wrapper" role="presentation" width="100%" cellpadding="0" cellspacing="0">
      <tr>
        <td align="center">
          <table class="container" role="presentation" cellpadding="0" cellspacing="0">
            <tr>
              <td class="divider">&nbsp;</td>
            </tr>
            <tr>
              <td class="header">
                <div class="logo">
                  <img src="brand_orange_bg_transparent@2xSite.svg" alt="Globalite">
                </div>
                <p class="eyebrow">{eyebrow}</p>
                <h1>{title}</h1>
                <p class="subtitle">{subtitle}</p>
                <p class="block-height">This article was written at block height: <strong>{block_height}</strong></p>
              </td>
            </tr>
{points_html}
            <tr>
              <td class="section tldr">
                <h2>{tldr_title}</h2>
{tldr_content}
              </td>
            </tr>
            <tr>
              <td class="section conclusion">
                <h2>{conclusion_title}</h2>
{conclusion_content}
              </td>
            </tr>
{snapshot_html}
            <tr>
              <td class="footer">
                <p>{footer_line}</p>
                <p>{address_line}</p>
                <div class="footer-links">
                  <a class="footer-logo-link" href="https://globalite.co" target="_blank" rel="noopener noreferrer">
                    <img src="{footer_logo_url}" alt="Globalite logo">
                  </a>
                  <div class="footer-social">
                    <a href="https://www.instagram.com/globalite.sa/" target="_blank" rel="noopener noreferrer"><img src="{footer_instagram_icon}" alt="Instagram"></a>
                    <a href="https://x.com/globalite_sa" target="_blank" rel="noopener noreferrer"><img src="{footer_x_icon}" alt="X"></a>
                    <a href="https://www.linkedin.com/company/globalite-sa" target="_blank" rel="noopener noreferrer"><img src="{footer_linkedin_icon}" alt="LinkedIn"></a>
                  </div>
                </div>
              </td>
            </tr>
          </table>
        </td>
      </tr>
    </table>
    <script>
      (function () {{
        var params = new URLSearchParams(window.location.search);
        var refreshSeconds = Number(params.get("refresh"));
        if (!Number.isFinite(refreshSeconds) || refreshSeconds < 5) {{
          return;
        }}
        window.setInterval(function () {{
          window.location.reload();
        }}, refreshSeconds * 1000);
      }})();
    </script>
  </body>
</html>
"""


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Generate newsletter HTML from Excel or Google Sheets."
    )
    parser.add_argument(
        "--xlsx",
        default="newsletter_data.xlsx",
        help="Input workbook path (default: newsletter_data.xlsx). Ignored when --google-sheet is set.",
    )
    parser.add_argument(
        "--google-sheet",
        default="",
        help="Google Sheet URL or ID. When set, this is used as the source instead of --xlsx.",
    )
    parser.add_argument(
        "--google-meta-tab",
        default="meta",
        help="Tab name for meta values in Google Sheets (default: meta).",
    )
    parser.add_argument(
        "--google-points-tab",
        default="points",
        help="Tab name for points in Google Sheets (default: points).",
    )
    parser.add_argument(
        "--google-distribution-tab",
        default="distribution",
        help="Tab name for distribution in Google Sheets (default: distribution). Optional.",
    )
    parser.add_argument(
        "--out",
        default="newsletter.html",
        help="Output HTML file path (default: newsletter.html)",
    )
    parser.add_argument(
        "--init-template",
        action="store_true",
        help="Create a starter workbook template and exit.",
    )
    parser.add_argument(
        "--force",
        action="store_true",
        help="Overwrite files when used with --init-template.",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    base_dir = Path(__file__).resolve().parent
    xlsx_path = Path(args.xlsx)
    if not xlsx_path.is_absolute():
        xlsx_path = base_dir / xlsx_path
    out_path = Path(args.out)
    if not out_path.is_absolute():
        out_path = base_dir / out_path
    google_sheet_ref = normalize_text(args.google_sheet)

    if args.init_template:
        if google_sheet_ref:
            raise ValueError("--init-template only works with local Excel files.")
        xlsx_path.parent.mkdir(parents=True, exist_ok=True)
        create_template_workbook(xlsx_path, force=args.force)
        print(f"Template created: {xlsx_path}")
        return 0

    meta: dict[str, str]
    points: list[Point]
    distribution: list[DistributionSegment]
    backup_path: Path

    if google_sheet_ref:
        sheet_id = extract_google_sheet_id(google_sheet_ref)
        meta_rows = fetch_google_sheet_rows(sheet_id, args.google_meta_tab, required=True)
        points_rows = fetch_google_sheet_rows(
            sheet_id, args.google_points_tab, required=True
        )
        distribution_rows = fetch_google_sheet_rows(
            sheet_id, args.google_distribution_tab, required=False
        )

        history_dir = base_dir / "history"
        history_dir.mkdir(exist_ok=True)
        backup_path = next_history_workbook_path(history_dir)
        write_google_snapshot_workbook(
            backup_path, meta_rows, points_rows, distribution_rows
        )

        meta = read_meta(TabularSheet(meta_rows))
        points = read_points(TabularSheet(points_rows))
        max_supply_btc = parse_number(meta["max_supply_btc"], default=21_000_000)
        if distribution_rows:
            distribution = read_distribution(
                TabularSheet(distribution_rows), max_supply_btc
            )
        else:
            distribution = default_distribution_segments(max_supply_btc)
    else:
        if not xlsx_path.exists():
            raise FileNotFoundError(
                f"Workbook not found: {xlsx_path}. "
                "Run with --init-template first to create it."
            )

        history_dir = xlsx_path.parent / "history"
        history_dir.mkdir(exist_ok=True)
        backup_path = next_history_workbook_path(history_dir)
        shutil.copy2(xlsx_path, backup_path)

        workbook = load_workbook(xlsx_path, data_only=True)
        if "meta" not in workbook.sheetnames:
            raise ValueError("Workbook is missing required sheet: meta")
        if "points" not in workbook.sheetnames:
            raise ValueError("Workbook is missing required sheet: points")

        meta = read_meta(workbook["meta"])
        points = read_points(workbook["points"])
        max_supply_btc = parse_number(meta["max_supply_btc"], default=21_000_000)
        if "distribution" in workbook.sheetnames:
            distribution = read_distribution(workbook["distribution"], max_supply_btc)
        else:
            distribution = default_distribution_segments(max_supply_btc)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    html_output = render_html(meta, points, distribution, out_path.parent)

    out_path.write_text(html_output, encoding="utf-8")
    print(
        f"Generated {out_path} with {len(points)} points "
        f"(max allowed: {MAX_POINTS})."
    )
    print(f"Source snapshot saved: {backup_path}")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as error:  # pragma: no cover
        print(f"Error: {error}", file=sys.stderr)
        raise SystemExit(1)
